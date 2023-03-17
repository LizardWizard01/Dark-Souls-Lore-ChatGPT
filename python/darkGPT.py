import openpyxl
from openpyxl import load_workbook

wb = openpyxl.load_workbook(filename='dsexcel.xlsx')
print(type(wb))
wb.sheetnames
print(wb.sheetnames)
sheet = wb['Items']
# sheet[]
# I'm getting "list object is not callable" as an error here, and not sure what it is talking about.
# ebb: You needed to import what I added in line 2: the load_workbook.
tuple(sheet['A1':'M850'])
for cellObj in sheet.columns[5]:
    print(cellObj.value)
# We want columns B, E, and H
# print(tuple(sheet))